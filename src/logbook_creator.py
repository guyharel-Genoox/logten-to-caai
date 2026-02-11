#!/usr/bin/env python3
"""
Create a comprehensive Excel flight logbook from Coradine LogTen Pro export.

Reads a tab-delimited LogTen Pro export file and creates a multi-sheet
Excel workbook with flight log, summary, aircraft list, and milestones.

Usage:
    python -m src.logbook_creator --input export.txt --output logbook.xlsx
    python -m src.logbook_creator --input export.txt --output logbook.xlsx --pilot "John Doe"
"""

import csv
import re
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


# ============ Styles ============

HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
DATA_FONT = Font(name='Calibri', size=9)
SOLO_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
CHECKRIDE_FILL = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)

# Column definitions: (name, width)
MAIN_COLUMNS = [
    ("No.", 5), ("Date", 11), ("From", 7), ("To", 7), ("Route", 15),
    ("Aircraft Reg.", 12), ("Aircraft Type", 10), ("Make", 14), ("Model", 22),
    ("Engine Type", 12), ("Category", 10), ("Class", 16), ("Total Time", 10),
    ("PIC", 8), ("SIC", 8), ("Night", 8), ("Cross Country", 10),
    ("Actual Inst.", 10), ("Sim. Inst.", 10), ("Dual Received", 10),
    ("Dual Given", 10), ("Solo", 8), ("Simulator", 8), ("Multi-Pilot", 10),
    ("Day LDG", 8), ("Day T/O", 8), ("Night LDG", 8), ("Night T/O", 8),
    ("Approach 1", 18), ("Approach 2", 18), ("Approach 3", 18), ("Approach 4", 18),
    ("Holds", 7), ("IFR", 6), ("Go Arounds", 8), ("PIC Name", 18),
    ("Instructor", 22), ("Student", 15), ("Observer/Safety Pilot", 22),
    ("DPE/Examiner", 18), ("Distance (NM)", 10), ("Remarks", 50),
    ("Complex", 7), ("High Perf.", 7), ("EFIS", 6), ("Retractable", 8),
    ("Pressurized", 8), ("Review/IPC", 7),
]

# LogTen Pro field name mappings (indexed by column position)
FIELD_MAP = [
    None,  # No. (auto-generated)
    "flight_flightDate", "flight_from", "flight_to", "flight_route",
    "aircraft_aircraftID", "aircraftType_type", "aircraftType_make",
    "aircraftType_model", "aircraftType_selectedEngineType",
    "aircraftType_selectedCategory", "aircraftType_selectedAircraftClass",
    "flight_totalTime", "flight_pic", "flight_sic", "flight_night",
    "flight_crossCountry", "flight_actualInstrument", "flight_simulatedInstrument",
    "flight_dualReceived", "flight_dualGiven", "flight_solo", "flight_simulator",
    "flight_multiPilot", "flight_dayLandings", "flight_dayTakeoffs",
    "flight_nightLandings", "flight_nightTakeoffs",
    "flight_selectedApproach1", "flight_selectedApproach2",
    "flight_selectedApproach3", "flight_selectedApproach4",
    "flight_holds", "flight_ifr", "flight_goArounds",
    "flight_selectedCrewPIC", "flight_selectedCrewInstructor",
    "flight_selectedCrewStudent", "flight_selectedCrewObserver",
    None,  # DPE - derived from observer field
    "flight_distance", "flight_remarks",
    "aircraft_complex", "aircraft_highPerformance", "aircraft_efis",
    "aircraft_undercarriageRetractable", "aircraft_pressurized",
    None,  # Review/IPC
]

# Text columns that should not be auto-converted to numbers
TEXT_COLUMNS = {1, 4, 5, 6, 7, 8, 9, 10, 11, 28, 29, 30, 31, 35, 36, 37, 38, 39, 41}


def time_to_decimal(time_str):
    """Convert H:MM or HH:MM to decimal hours."""
    if not time_str or time_str.strip() == '':
        return None
    time_str = time_str.strip()
    match = re.match(r'^(\d+):(\d{2})$', time_str)
    if match:
        h, m = int(match.group(1)), int(match.group(2))
        return round(h + m / 60, 2)
    return None


def parse_logten_export(input_file):
    """Parse a LogTen Pro tab-delimited export file.

    Args:
        input_file: Path to the LogTen Pro export file (.txt, tab-delimited)

    Returns:
        Tuple of (col_map dict, data_rows list of raw strings)
    """
    with open(input_file, 'r', encoding='utf-8') as f:
        raw_content = f.read()

    lines = raw_content.split('\n')
    header_line = lines[0].strip()
    headers = [h.strip() for h in header_line.split('\t')]
    col_map = {h: i for i, h in enumerate(headers)}

    # Parse data rows handling multiline remarks
    data_rows = []
    current_row = None
    for i in range(1, len(lines)):
        line = lines[i]
        if not line.strip():
            continue
        if current_row is not None:
            first_field = line.split('\t')[0].strip()
            if re.match(r'^\d{4}-\d{2}-\d{2}$', first_field):
                data_rows.append(current_row)
                current_row = line
            else:
                current_row += ' ' + line
        else:
            current_row = line
    if current_row:
        data_rows.append(current_row)

    return col_map, data_rows


def create_logbook(input_file, output_file, pilot_name=""):
    """Create a comprehensive Excel flight logbook from LogTen Pro export.

    Args:
        input_file: Path to the LogTen Pro export file (.txt, tab-delimited)
        output_file: Path to the output Excel file (.xlsx)
        pilot_name: Pilot's name for the summary sheet title

    Returns:
        Dict with summary statistics.
    """
    col_map, data_rows = parse_logten_export(input_file)
    print(f"Parsed {len(data_rows)} flight records")

    def get_field(fields, field_name):
        idx = col_map.get(field_name)
        if idx is not None and idx < len(fields):
            val = fields[idx].strip().strip('"')
            val = re.sub(r'\s+', ' ', val)
            return val
        return ''

    # Create workbook
    wb = Workbook()

    # ============ SHEET 1: FLIGHT LOG ============
    ws = wb.active
    ws.title = "Flight Log"

    # Write header row
    for col_idx, (col_name, col_width) in enumerate(MAIN_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(MAIN_COLUMNS))}1"

    # Write data rows
    for row_num, row_str in enumerate(data_rows, 2):
        fields = row_str.split('\t')

        for col_idx in range(len(MAIN_COLUMNS)):
            sheet_col = col_idx + 1

            if col_idx == 0:  # Row number
                val = row_num - 1
            elif col_idx == 39:  # DPE/Examiner
                obs = get_field(fields, "flight_selectedCrewObserver")
                val = obs if 'DPE' in obs.upper() else ''
            elif col_idx == 47:  # Review/IPC
                review = get_field(fields, "flight_review")
                ipc = get_field(fields, "flight_instrumentProficiencyCheck")
                parts = []
                if review and review != '0':
                    parts.append("BFR")
                if ipc and ipc != '0':
                    parts.append("IPC")
                val = '/'.join(parts)
            elif FIELD_MAP[col_idx] is not None:
                val = get_field(fields, FIELD_MAP[col_idx])
            else:
                val = ''

            # Convert time values to decimal
            if col_idx in range(12, 25) and isinstance(val, str) and ':' in val:
                dec = time_to_decimal(val)
                if dec is not None:
                    val = dec

            # Convert numeric strings
            if isinstance(val, str) and val:
                try:
                    if '.' in val and col_idx not in TEXT_COLUMNS:
                        val = float(val)
                    elif val.isdigit() and col_idx not in TEXT_COLUMNS:
                        val = int(val)
                except (ValueError, AttributeError):
                    pass

            # Handle date column
            if col_idx == 1 and isinstance(val, str) and re.match(r'^\d{4}-\d{2}-\d{2}$', val):
                try:
                    val = datetime.strptime(val, '%Y-%m-%d')
                except ValueError:
                    pass

            cell = ws.cell(row=row_num, column=sheet_col, value=val if val != '' else None)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=(col_idx in [41, 28, 29, 30, 31]))

            if col_idx == 1 and isinstance(val, datetime):
                cell.number_format = 'YYYY-MM-DD'
            if col_idx in range(12, 25) and isinstance(val, (int, float)):
                cell.number_format = '0.00'

        # Highlight special rows
        remarks = get_field(fields, "flight_remarks")
        solo_val = get_field(fields, "flight_solo")

        if 'Solo' in remarks or (solo_val and solo_val not in ['', '0']):
            for c in range(1, len(MAIN_COLUMNS) + 1):
                ws.cell(row=row_num, column=c).fill = SOLO_FILL

        if 'checkride' in remarks.lower():
            for c in range(1, len(MAIN_COLUMNS) + 1):
                ws.cell(row=row_num, column=c).fill = CHECKRIDE_FILL

    # ============ SHEET 2: SUMMARY ============
    ws2 = wb.create_sheet("Summary & Totals")

    # Calculate totals
    totals = {
        'total_flights': len(data_rows),
        'total_time': 0, 'pic_time': 0, 'sic_time': 0, 'night_time': 0,
        'xc_time': 0, 'actual_inst': 0, 'sim_inst': 0, 'dual_recv': 0,
        'dual_given': 0, 'solo_time': 0, 'sim_time': 0, 'multi_pilot': 0,
        'day_ldg': 0, 'day_to': 0, 'night_ldg': 0, 'night_to': 0,
        'holds': 0, 'approaches': 0,
    }
    aircraft_types = {}
    airports_from = {}
    yearly_hours = {}
    instructors = {}

    for row_str in data_rows:
        fields = row_str.split('\t')
        for key, field in [
            ('total_time', 'flight_totalTime'), ('pic_time', 'flight_pic'),
            ('sic_time', 'flight_sic'), ('night_time', 'flight_night'),
            ('xc_time', 'flight_crossCountry'), ('actual_inst', 'flight_actualInstrument'),
            ('sim_inst', 'flight_simulatedInstrument'), ('dual_recv', 'flight_dualReceived'),
            ('dual_given', 'flight_dualGiven'), ('solo_time', 'flight_solo'),
            ('sim_time', 'flight_simulator'), ('multi_pilot', 'flight_multiPilot'),
        ]:
            dec = time_to_decimal(get_field(fields, field))
            if dec:
                totals[key] += dec

        for key, field in [
            ('day_ldg', 'flight_dayLandings'), ('day_to', 'flight_dayTakeoffs'),
            ('night_ldg', 'flight_nightLandings'), ('night_to', 'flight_nightTakeoffs'),
        ]:
            val = get_field(fields, field)
            if val and val.isdigit():
                totals[key] += int(val)

        val = get_field(fields, 'flight_holds')
        if val and val.isdigit():
            totals['holds'] += int(val)

        for ap_field in ['flight_selectedApproach1', 'flight_selectedApproach2',
                         'flight_selectedApproach3', 'flight_selectedApproach4']:
            if get_field(fields, ap_field):
                totals['approaches'] += 1

        atype = get_field(fields, 'aircraftType_type')
        if atype:
            tt = time_to_decimal(get_field(fields, 'flight_totalTime')) or 0
            aircraft_types[atype] = aircraft_types.get(atype, {'flights': 0, 'hours': 0})
            aircraft_types[atype]['flights'] += 1
            aircraft_types[atype]['hours'] += tt

        afrom = get_field(fields, 'flight_from')
        if afrom:
            airports_from[afrom] = airports_from.get(afrom, 0) + 1

        date_str = get_field(fields, 'flight_flightDate')
        if date_str and len(date_str) >= 4:
            year = date_str[:4]
            tt = time_to_decimal(get_field(fields, 'flight_totalTime')) or 0
            yearly_hours[year] = yearly_hours.get(year, {'flights': 0, 'hours': 0})
            yearly_hours[year]['flights'] += 1
            yearly_hours[year]['hours'] += tt

        instr = get_field(fields, 'flight_selectedCrewInstructor')
        if instr:
            instructors[instr] = instructors.get(instr, 0) + 1

    # Write Summary
    title_font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
    section_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    section_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    label_font = Font(name='Calibri', bold=True, size=10)
    value_font = Font(name='Calibri', size=10)

    for col_letter, width in [('A', 25), ('B', 15), ('C', 5), ('D', 25), ('E', 15)]:
        ws2.column_dimensions[col_letter].width = width

    title = f"{pilot_name} - Flight Logbook Summary" if pilot_name else "Flight Logbook Summary"
    ws2.cell(row=1, column=1, value=title).font = title_font
    ws2.merge_cells('A1:E1')

    # Career Totals
    row = 3
    for col in range(1, 6):
        ws2.cell(row=row, column=col).fill = section_fill
    ws2.cell(row=row, column=1, value="CAREER TOTALS").font = section_font

    row = 4
    summary_items = [
        ("Total Flights", totals['total_flights']),
        ("Total Flight Time", f"{totals['total_time']:.1f} hrs"),
        ("PIC Time", f"{totals['pic_time']:.1f} hrs"),
        ("SIC Time", f"{totals['sic_time']:.1f} hrs"),
        ("Night Time", f"{totals['night_time']:.1f} hrs"),
        ("Cross Country", f"{totals['xc_time']:.1f} hrs"),
        ("Actual Instrument", f"{totals['actual_inst']:.1f} hrs"),
        ("Simulated Instrument", f"{totals['sim_inst']:.1f} hrs"),
        ("Dual Received", f"{totals['dual_recv']:.1f} hrs"),
        ("Dual Given", f"{totals['dual_given']:.1f} hrs"),
        ("Solo", f"{totals['solo_time']:.1f} hrs"),
        ("Simulator", f"{totals['sim_time']:.1f} hrs"),
        ("Multi-Pilot", f"{totals['multi_pilot']:.1f} hrs"),
        ("", ""),
        ("Day Landings", totals['day_ldg']),
        ("Day Takeoffs", totals['day_to']),
        ("Night Landings", totals['night_ldg']),
        ("Night Takeoffs", totals['night_to']),
        ("Total Landings", totals['day_ldg'] + totals['night_ldg']),
        ("Total Takeoffs", totals['day_to'] + totals['night_to']),
        ("", ""),
        ("Instrument Approaches", totals['approaches']),
        ("Holds", totals['holds']),
    ]

    for label, value in summary_items:
        if label:
            ws2.cell(row=row, column=1, value=label).font = label_font
            ws2.cell(row=row, column=2, value=value).font = value_font
            ws2.cell(row=row, column=1).border = THIN_BORDER
            ws2.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    # Yearly Breakdown
    row += 1
    for col in range(1, 6):
        ws2.cell(row=row, column=col).fill = section_fill
    ws2.cell(row=row, column=1, value="YEARLY BREAKDOWN").font = section_font
    row += 1

    for hdr, c in [("Year", 1), ("Flights", 2), ("Hours", 4)]:
        cell = ws2.cell(row=row, column=c, value=hdr)
        cell.fill = SUBHEADER_FILL
        cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        cell.border = THIN_BORDER
    row += 1

    for year in sorted(yearly_hours.keys()):
        ws2.cell(row=row, column=1, value=year).font = value_font
        ws2.cell(row=row, column=2, value=yearly_hours[year]['flights']).font = value_font
        ws2.cell(row=row, column=4, value=f"{yearly_hours[year]['hours']:.1f}").font = value_font
        for c in [1, 2, 4]:
            ws2.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # Aircraft Types
    row += 1
    for col in range(1, 6):
        ws2.cell(row=row, column=col).fill = section_fill
    ws2.cell(row=row, column=1, value="AIRCRAFT TYPES FLOWN").font = section_font
    row += 1

    for hdr, c in [("Type", 1), ("Flights", 2), ("Hours", 4)]:
        cell = ws2.cell(row=row, column=c, value=hdr)
        cell.fill = SUBHEADER_FILL
        cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        cell.border = THIN_BORDER
    row += 1

    for atype in sorted(aircraft_types.keys(), key=lambda x: aircraft_types[x]['hours'], reverse=True):
        ws2.cell(row=row, column=1, value=atype).font = value_font
        ws2.cell(row=row, column=2, value=aircraft_types[atype]['flights']).font = value_font
        ws2.cell(row=row, column=4, value=f"{aircraft_types[atype]['hours']:.1f}").font = value_font
        for c in [1, 2, 4]:
            ws2.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # Top Airports
    row += 1
    for col in range(1, 6):
        ws2.cell(row=row, column=col).fill = section_fill
    ws2.cell(row=row, column=1, value="TOP AIRPORTS (DEPARTURES)").font = section_font
    row += 1

    for hdr, c in [("Airport", 1), ("Departures", 2)]:
        cell = ws2.cell(row=row, column=c, value=hdr)
        cell.fill = SUBHEADER_FILL
        cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        cell.border = THIN_BORDER
    row += 1

    for airport in sorted(airports_from.keys(), key=lambda x: airports_from[x], reverse=True)[:20]:
        ws2.cell(row=row, column=1, value=airport).font = value_font
        ws2.cell(row=row, column=2, value=airports_from[airport]).font = value_font
        for c in [1, 2]:
            ws2.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # Instructors
    row += 1
    for col in range(1, 6):
        ws2.cell(row=row, column=col).fill = section_fill
    ws2.cell(row=row, column=1, value="INSTRUCTORS TRAINED WITH").font = section_font
    row += 1

    for hdr, c in [("Instructor", 1), ("Flights", 2)]:
        cell = ws2.cell(row=row, column=c, value=hdr)
        cell.fill = SUBHEADER_FILL
        cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        cell.border = THIN_BORDER
    row += 1

    for instr in sorted(instructors.keys(), key=lambda x: instructors[x], reverse=True):
        ws2.cell(row=row, column=1, value=instr).font = value_font
        ws2.cell(row=row, column=2, value=instructors[instr]).font = value_font
        for c in [1, 2]:
            ws2.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # ============ SHEET 3: AIRCRAFT LIST ============
    ws3 = wb.create_sheet("Aircraft")
    aircraft_list = {}
    for row_str in data_rows:
        fields = row_str.split('\t')
        reg = get_field(fields, 'aircraft_aircraftID')
        if reg and reg not in aircraft_list:
            aircraft_list[reg] = {
                'type': get_field(fields, 'aircraftType_type'),
                'make': get_field(fields, 'aircraftType_make'),
                'model': get_field(fields, 'aircraftType_model'),
                'engine': get_field(fields, 'aircraftType_selectedEngineType'),
                'category': get_field(fields, 'aircraftType_selectedCategory'),
                'class': get_field(fields, 'aircraftType_selectedAircraftClass'),
                'complex': get_field(fields, 'aircraft_complex'),
                'hp': get_field(fields, 'aircraft_highPerformance'),
                'efis': get_field(fields, 'aircraft_efis'),
                'retract': get_field(fields, 'aircraft_undercarriageRetractable'),
                'press': get_field(fields, 'aircraft_pressurized'),
            }

    ac_headers = ["Registration", "Type Code", "Make", "Model", "Engine Type",
                  "Category", "Class", "Complex", "High Perf.", "EFIS",
                  "Retractable", "Pressurized"]
    ac_widths = [14, 12, 25, 40, 14, 12, 18, 9, 9, 7, 10, 10]

    for col_idx, (hdr, w) in enumerate(zip(ac_headers, ac_widths), 1):
        cell = ws3.cell(row=1, column=col_idx, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
        ws3.column_dimensions[get_column_letter(col_idx)].width = w
    ws3.freeze_panes = 'A2'

    row_num = 2
    for reg in sorted(aircraft_list.keys()):
        ac = aircraft_list[reg]
        vals = [reg, ac['type'], ac['make'], ac['model'], ac['engine'],
                ac['category'], ac['class'], ac['complex'], ac['hp'],
                ac['efis'], ac['retract'], ac['press']]
        for col_idx, val in enumerate(vals, 1):
            cell = ws3.cell(row=row_num, column=col_idx, value=val if val else None)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row_num += 1

    # ============ SHEET 4: MILESTONES ============
    ws4 = wb.create_sheet("Milestones")
    milestone_headers = ["Date", "Event", "Details", "Aircraft", "Location", "Examiner/Instructor"]
    milestone_widths = [12, 25, 50, 12, 8, 25]

    for col_idx, (hdr, w) in enumerate(zip(milestone_headers, milestone_widths), 1):
        cell = ws4.cell(row=1, column=col_idx, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws4.column_dimensions[get_column_letter(col_idx)].width = w
    ws4.freeze_panes = 'A2'

    milestones = []
    for row_str in data_rows:
        fields = row_str.split('\t')
        remarks = get_field(fields, 'flight_remarks')
        date = get_field(fields, 'flight_flightDate')
        aircraft = get_field(fields, 'aircraft_aircraftID')
        loc_from = get_field(fields, 'flight_from')
        observer = get_field(fields, 'flight_selectedCrewObserver')
        instructor = get_field(fields, 'flight_selectedCrewInstructor')

        remarks_lower = remarks.lower()
        event = None
        if 'first solo' in remarks_lower:
            event = "First Solo"
        elif 'ppl checkride' in remarks_lower or 'complete ppl' in remarks_lower:
            event = "PPL Checkride"
        elif 'ir checkride' in remarks_lower:
            event = "IR Checkride"
        elif 'cpl checkride single' in remarks_lower:
            event = "CPL Single Add-on Checkride"
        elif 'cpl checkride' in remarks_lower:
            event = "CPL Checkride"
        elif 'cfi checkride' in remarks_lower or 'cfi initial' in remarks_lower:
            event = "CFI Checkride"
        elif 'cfii checkride' in remarks_lower:
            event = "CFII Checkride"
        elif 'mei checkride' in remarks_lower:
            event = "MEI Checkride"
        elif 'atp' in remarks_lower and 'checkride' in remarks_lower:
            event = "ATP Checkride"
        elif 'type rating' in remarks_lower:
            event = "Type Rating"
        elif 'spin training' in remarks_lower:
            event = "Spin Training & Endorsement"
        elif 'checkout' in remarks_lower or 'check out' in remarks_lower:
            event = "Aircraft Checkout"

        if event:
            examiner = observer if 'DPE' in observer.upper() else instructor
            milestones.append((date, event, remarks, aircraft, loc_from, examiner))

    for row_idx, (date, event, details, aircraft, loc, examiner) in enumerate(milestones, 2):
        vals = [date, event, details, aircraft, loc, examiner]
        for col_idx, val in enumerate(vals, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center' if col_idx != 3 else 'left',
                                       vertical='center', wrap_text=(col_idx == 3))
            if 'Checkride' in event:
                cell.fill = CHECKRIDE_FILL

    # Save
    wb.save(output_file)
    print(f"\nExcel file created: {output_file}")
    print(f"Sheets: Flight Log | Summary & Totals | Aircraft | Milestones")
    print(f"\nSummary:")
    print(f"  Total Flights: {totals['total_flights']}")
    print(f"  Total Time: {totals['total_time']:.1f} hrs")
    print(f"  PIC Time: {totals['pic_time']:.1f} hrs")
    print(f"  Aircraft Types: {len(aircraft_types)}")
    print(f"  Unique Aircraft: {len(aircraft_list)}")

    return totals


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Create Excel logbook from LogTen Pro export')
    parser.add_argument('--input', '-i', required=True, help='LogTen Pro tab-delimited export file')
    parser.add_argument('--output', '-o', required=True, help='Output Excel file path')
    parser.add_argument('--pilot', '-p', default='', help='Pilot name for summary sheet title')
    args = parser.parse_args()
    create_logbook(args.input, args.output, args.pilot)
