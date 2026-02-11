#!/usr/bin/env python3
"""
Analyze flight logbook data to categorize per CAAI (Israeli CAA) rules.

Outputs categorized totals for each aircraft type and verifies CAAI
rule compliance. Useful for verification and debugging.

Usage:
    python -m src.caai_analyzer --logbook logbook.xlsx
"""

import argparse
from openpyxl import load_workbook
from collections import defaultdict

from .caai_rules import (
    is_simulator, get_caai_category, is_single_engine, normalize_type,
    CAAI_GROUP_MAP,
)
from .column_map import (
    COL_DATE, COL_FROM, COL_TO, COL_REGISTRATION, COL_AIRCRAFT_TYPE,
    COL_MAKE, COL_MODEL, COL_ENGINE_TYPE, COL_CATEGORY, COL_CLASS,
    COL_TOTAL_TIME, COL_PIC, COL_SIC, COL_NIGHT, COL_CROSS_COUNTRY,
    COL_ACTUAL_INSTRUMENT, COL_SIMULATED_INSTRUMENT, COL_DUAL_RECEIVED,
    COL_DUAL_GIVEN, COL_SOLO, COL_SIMULATOR, COL_MULTI_PILOT,
    COL_DAY_LANDINGS, COL_DAY_TAKEOFFS, COL_NIGHT_LANDINGS, COL_NIGHT_TAKEOFFS,
    COL_PIC_NAME, COL_INSTRUCTOR, COL_DISTANCE, COL_REMARKS,
)


def analyze_caai(logbook_file):
    """Analyze and categorize all flights per CAAI rules.

    Args:
        logbook_file: Path to the flight logbook Excel file.

    Returns:
        Dict with analysis results.
    """
    wb = load_workbook(logbook_file)
    ws = wb["Flight Log"]

    flights = []
    for row in range(2, ws.max_row + 1):
        f = {}
        f['row'] = row
        f['date'] = ws.cell(row=row, column=COL_DATE).value
        f['from'] = str(ws.cell(row=row, column=COL_FROM).value or '').strip()
        f['to'] = str(ws.cell(row=row, column=COL_TO).value or '').strip()
        f['reg'] = str(ws.cell(row=row, column=COL_REGISTRATION).value or '').strip()
        f['aircraft_type'] = str(ws.cell(row=row, column=COL_AIRCRAFT_TYPE).value or '').strip()
        f['make'] = str(ws.cell(row=row, column=COL_MAKE).value or '').strip()
        f['model'] = str(ws.cell(row=row, column=COL_MODEL).value or '').strip()
        f['engine'] = str(ws.cell(row=row, column=COL_ENGINE_TYPE).value or '').strip()
        f['category'] = str(ws.cell(row=row, column=COL_CATEGORY).value or '').strip()
        f['class'] = str(ws.cell(row=row, column=COL_CLASS).value or '').strip()
        f['total'] = float(ws.cell(row=row, column=COL_TOTAL_TIME).value or 0)
        f['pic'] = float(ws.cell(row=row, column=COL_PIC).value or 0)
        f['sic'] = float(ws.cell(row=row, column=COL_SIC).value or 0)
        f['night'] = float(ws.cell(row=row, column=COL_NIGHT).value or 0)
        f['xc'] = float(ws.cell(row=row, column=COL_CROSS_COUNTRY).value or 0)
        f['actual_inst'] = float(ws.cell(row=row, column=COL_ACTUAL_INSTRUMENT).value or 0)
        f['sim_inst'] = float(ws.cell(row=row, column=COL_SIMULATED_INSTRUMENT).value or 0)
        f['dual_recv'] = float(ws.cell(row=row, column=COL_DUAL_RECEIVED).value or 0)
        f['dual_given'] = float(ws.cell(row=row, column=COL_DUAL_GIVEN).value or 0)
        f['solo'] = float(ws.cell(row=row, column=COL_SOLO).value or 0)
        f['simulator'] = float(ws.cell(row=row, column=COL_SIMULATOR).value or 0)
        f['multi_pilot'] = float(ws.cell(row=row, column=COL_MULTI_PILOT).value or 0)
        f['day_ldg'] = int(ws.cell(row=row, column=COL_DAY_LANDINGS).value or 0)
        f['day_to'] = int(ws.cell(row=row, column=COL_DAY_TAKEOFFS).value or 0)
        f['night_ldg'] = int(ws.cell(row=row, column=COL_NIGHT_LANDINGS).value or 0)
        f['night_to'] = int(ws.cell(row=row, column=COL_NIGHT_TAKEOFFS).value or 0)
        f['pic_name'] = str(ws.cell(row=row, column=COL_PIC_NAME).value or '').strip()
        f['instructor'] = str(ws.cell(row=row, column=COL_INSTRUCTOR).value or '').strip()
        f['remarks'] = str(ws.cell(row=row, column=COL_REMARKS).value or '').strip()
        dist_val = ws.cell(row=row, column=COL_DISTANCE).value or 0
        if isinstance(dist_val, str):
            dist_val = dist_val.replace(',', '')
        f['distance'] = float(dist_val)
        flights.append(f)

    wb.close()

    # Categorize
    type_totals = defaultdict(lambda: {
        'category': '', 'is_sim': False, 'is_multi': False, 'total': 0,
        'day_pic': 0, 'day_pic_xc': 0, 'day_sic': 0, 'day_student': 0,
        'night_pic': 0, 'night_pic_xc': 0, 'night_sic': 0, 'night_student': 0,
        'actual_inst': 0, 'sim_inst_air': 0, 'sim_time': 0,
        'day_ldg': 0, 'night_ldg': 0, 'flights': 0,
    })

    grand_totals = {
        'total_flights': 0, 'total_time': 0,
        'pic_time': 0, 'sic_time': 0, 'student_time': 0, 'solo_time': 0,
        'instructor_time': 0,
        'night_time': 0, 'xc_time': 0,
        'actual_inst': 0, 'sim_inst': 0, 'sim_time': 0,
        'day_ldg': 0, 'night_ldg': 0,
        'pic_xc': 0, 'night_pic': 0, 'night_pic_xc': 0,
        'dual_night': 0, 'solo_xc': 0,
        'safety_pilot_se': 0,
    }

    for f in flights:
        atype = f['aircraft_type']
        if not atype:
            continue

        sim = is_simulator(atype, f['reg'])
        single_engine = is_single_engine(atype)
        t = type_totals[atype]
        t['is_sim'] = sim
        t['is_multi'] = not single_engine
        t['flights'] += 1

        if sim:
            t['sim_time'] += f['total']
            t['category'] = 'simulator'
            grand_totals['sim_time'] += f['total']
            t['actual_inst'] += f['actual_inst']
            t['sim_inst_air'] += f['sim_inst']
            continue

        caai_cat = get_caai_category(atype)
        t['category'] = caai_cat
        t['total'] += f['total']
        grand_totals['total_flights'] += 1
        grand_totals['total_time'] += f['total']

        day_time = f['total'] - f['night']
        night_time = f['night']

        grand_totals['night_time'] += night_time
        grand_totals['actual_inst'] += f['actual_inst']
        grand_totals['sim_inst'] += f['sim_inst']
        grand_totals['day_ldg'] += f['day_ldg']
        grand_totals['night_ldg'] += f['night_ldg']

        t['actual_inst'] += f['actual_inst']
        t['sim_inst_air'] += f['sim_inst']
        t['day_ldg'] += f['day_ldg']
        t['night_ldg'] += f['night_ldg']

        has_instructor = bool(f['instructor']) or f['dual_recv'] > 0
        is_instructor = f['dual_given'] > 0
        is_solo = f['solo'] > 0
        is_sic = f['sic'] > 0 and not has_instructor
        is_safety = 'safety pilot' in f['remarks'].lower()
        is_xc = f['xc'] > 0 or f['distance'] > 27

        if is_xc:
            grand_totals['xc_time'] += f['total']

        if has_instructor:
            t['day_student'] += day_time
            t['night_student'] += night_time
            grand_totals['student_time'] += f['total']
            if night_time > 0:
                grand_totals['dual_night'] += night_time
        elif is_safety and single_engine:
            grand_totals['safety_pilot_se'] += f['total']
        elif is_instructor:
            t['day_pic'] += day_time
            t['night_pic'] += night_time
            if is_xc:
                t['day_pic_xc'] += day_time
                t['night_pic_xc'] += night_time
                grand_totals['pic_xc'] += f['total']
            grand_totals['pic_time'] += f['total']
            grand_totals['instructor_time'] += f['total']
            if night_time > 0:
                grand_totals['night_pic'] += night_time
                if is_xc:
                    grand_totals['night_pic_xc'] += night_time
        elif is_solo:
            t['day_pic'] += day_time
            t['night_pic'] += night_time
            if is_xc:
                t['day_pic_xc'] += day_time
                t['night_pic_xc'] += night_time
                grand_totals['pic_xc'] += f['total']
                grand_totals['solo_xc'] += f['total']
            grand_totals['pic_time'] += f['total']
            grand_totals['solo_time'] += f['total']
            if night_time > 0:
                grand_totals['night_pic'] += night_time
                if is_xc:
                    grand_totals['night_pic_xc'] += night_time
        elif is_sic and not single_engine:
            t['day_sic'] += day_time
            t['night_sic'] += night_time
            grand_totals['sic_time'] += f['total']
        elif f['pic'] > 0:
            t['day_pic'] += day_time
            t['night_pic'] += night_time
            if is_xc:
                t['day_pic_xc'] += day_time
                t['night_pic_xc'] += night_time
                grand_totals['pic_xc'] += f['total']
            grand_totals['pic_time'] += f['total']
            if night_time > 0:
                grand_totals['night_pic'] += night_time
                if is_xc:
                    grand_totals['night_pic_xc'] += night_time
        else:
            if is_safety:
                t['day_sic'] += day_time
                t['night_sic'] += night_time
                grand_totals['sic_time'] += f['total']
            elif f['total'] > 0:
                t['day_pic'] += day_time
                t['night_pic'] += night_time
                if is_xc:
                    t['day_pic_xc'] += day_time
                    t['night_pic_xc'] += night_time
                    grand_totals['pic_xc'] += f['total']
                grand_totals['pic_time'] += f['total']
                if night_time > 0:
                    grand_totals['night_pic'] += night_time
                    if is_xc:
                        grand_totals['night_pic_xc'] += night_time

    # Print results
    print("=" * 90)
    print("CAAI FORM DATA - CATEGORIZED BY AIRCRAFT TYPE")
    print("=" * 90)

    print(f"\n{'Type':<12} {'Cat':<4} {'Flt':>4} {'Total':>7} {'DayPIC':>7} {'DayPICxc':>8} {'DaySIC':>7} {'DayStd':>7} {'NtPIC':>7} {'NtSIC':>7} {'NtStd':>7} {'ActI':>6} {'SimI':>6}")
    print("-" * 90)

    for atype in sorted(type_totals.keys(), key=lambda x: type_totals[x]['total'], reverse=True):
        t = type_totals[atype]
        if t['is_sim']:
            print(f"{atype:<12} {'SIM':<4} {t['flights']:>4} {t['sim_time']:>7.1f} {'---SIM---':>56}")
        else:
            print(f"{atype:<12} {t['category']:<4} {t['flights']:>4} {t['total']:>7.1f} {t['day_pic']:>7.1f} {t['day_pic_xc']:>8.1f} {t['day_sic']:>7.1f} {t['day_student']:>7.1f} {t['night_pic']:>7.1f} {t['night_sic']:>7.1f} {t['night_student']:>7.1f} {t['actual_inst']:>6.1f} {t['sim_inst_air']:>6.1f}")

    print("\n" + "=" * 90)
    print("GRAND TOTALS (aircraft only, excludes simulators)")
    print("=" * 90)
    for k, v in sorted(grand_totals.items()):
        print(f"  {k:<20}: {v:.1f}" if isinstance(v, float) else f"  {k:<20}: {v}")

    # Per-category totals
    print("\n" + "=" * 90)
    print("PER CAAI CATEGORY TOTALS (for Table 1)")
    print("=" * 90)

    cat_totals = defaultdict(lambda: defaultdict(float))
    for atype, t in type_totals.items():
        if t['is_sim']:
            continue
        cat = t['category']
        for key in ['total', 'day_pic', 'day_pic_xc', 'day_sic', 'day_student',
                     'night_pic', 'night_pic_xc', 'night_sic', 'night_student',
                     'actual_inst', 'sim_inst_air']:
            cat_totals[cat][key] += t[key]

    for cat in sorted(cat_totals.keys()):
        ct = cat_totals[cat]
        group_letter = CAAI_GROUP_MAP.get(cat, cat)
        print(f"\n  {cat} ({group_letter}):")
        for k, v in ct.items():
            if v > 0:
                print(f"    {k:<20}: {v:.1f}")

    # SIC half-credit
    sic_half = grand_totals['sic_time'] / 2
    caai_total = grand_totals['pic_time'] + grand_totals['student_time'] + sic_half
    print(f"\n\nSIC Half-Credit (per regulation 42(b)): {sic_half:.1f} hrs")
    print(f"Total for CAAI (PIC + Student + SIC/2): {caai_total:.1f} hrs")
    print(f"Safety pilot on SE (excluded): {grand_totals['safety_pilot_se']:.1f} hrs")

    return {
        'type_totals': dict(type_totals),
        'grand_totals': grand_totals,
        'cat_totals': dict(cat_totals),
        'caai_total': caai_total,
    }


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Analyze logbook per CAAI rules')
    parser.add_argument('--logbook', '-l', required=True, help='Flight logbook Excel file')
    args = parser.parse_args()
    analyze_caai(args.logbook)
