#!/usr/bin/env python3
"""
Add great-circle distance (NM) to the flight logbook Excel.

Uses the airport coordinate database to calculate haversine distances
for each flight leg. Fills in column 41 (Distance NM).

Usage:
    python -m src.distance_calculator --logbook logbook.xlsx
"""

import argparse
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

from .airports import get_all_airports, haversine_nm, SIMULATOR_ENTRIES
from .column_map import COL_FROM, COL_TO, COL_DISTANCE


DATA_FONT = Font(name='Calibri', size=9)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)


def add_distances(logbook_file, custom_airports_file=None):
    """Calculate and add great-circle distances to the logbook.

    Fills column 41 (Distance NM) using haversine formula.
    Skips same-airport flights (pattern work, set to 0).
    Skips flights where departure or arrival airport is not in the database.

    Args:
        logbook_file: Path to the flight logbook Excel file.
        custom_airports_file: Optional path to JSON file with additional airports.

    Returns:
        Dict with stats: {filled, updated, skipped, not_found}.
    """
    airports = get_all_airports(custom_airports_file)

    wb = load_workbook(logbook_file)
    ws = wb["Flight Log"]

    print(f"Column {COL_DISTANCE} header: {ws.cell(row=1, column=COL_DISTANCE).value}")

    filled = 0
    updated = 0
    skipped = 0
    not_found = set()

    for row in range(2, ws.max_row + 1):
        from_apt = ws.cell(row=row, column=COL_FROM).value
        to_apt = ws.cell(row=row, column=COL_TO).value
        existing_dist = ws.cell(row=row, column=COL_DISTANCE).value

        if not from_apt or not to_apt:
            skipped += 1
            continue

        from_apt = str(from_apt).strip()
        to_apt = str(to_apt).strip()

        # Skip same-airport flights (pattern work)
        if from_apt == to_apt:
            if existing_dist is None:
                ws.cell(row=row, column=COL_DISTANCE).value = 0
                ws.cell(row=row, column=COL_DISTANCE).number_format = '0.0'
                filled += 1
            continue

        # Look up coordinates
        from_coords = airports.get(from_apt)
        to_coords = airports.get(to_apt)

        if from_coords is None or to_coords is None:
            if from_coords is None and from_apt not in airports and from_apt not in SIMULATOR_ENTRIES:
                not_found.add(from_apt)
            if to_coords is None and to_apt not in airports and to_apt not in SIMULATOR_ENTRIES:
                not_found.add(to_apt)
            skipped += 1
            continue

        # Calculate distance
        dist = haversine_nm(from_coords[0], from_coords[1], to_coords[0], to_coords[1])

        if existing_dist is None or existing_dist == '' or existing_dist == 0:
            cell = ws.cell(row=row, column=COL_DISTANCE)
            cell.value = dist
            cell.number_format = '0.0'
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            filled += 1
        else:
            try:
                existing_val = float(existing_dist)
                # Keep existing if similar (may reflect actual routing)
            except (ValueError, TypeError):
                ws.cell(row=row, column=COL_DISTANCE).value = dist
                ws.cell(row=row, column=COL_DISTANCE).number_format = '0.0'
                updated += 1

    if not_found:
        print(f"\nAirports not found in database ({len(not_found)}):")
        for a in sorted(not_found):
            print(f"  {a}")

    print(f"\nDistances filled: {filled}")
    print(f"Distances updated: {updated}")
    print(f"Skipped: {skipped}")

    wb.save(logbook_file)
    print(f"\nExcel file updated: {logbook_file}")

    return {'filled': filled, 'updated': updated, 'skipped': skipped,
            'not_found': sorted(not_found)}


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Add distances to flight logbook')
    parser.add_argument('--logbook', '-l', required=True, help='Flight logbook Excel file')
    parser.add_argument('--airports', '-a', default=None, help='Custom airports JSON file')
    args = parser.parse_args()
    add_distances(args.logbook, args.airports)
