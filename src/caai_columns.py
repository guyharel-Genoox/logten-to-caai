#!/usr/bin/env python3
"""
Add CAAI-required columns to the flight logbook Excel.

Adds columns 49-56 (AW-BD): CAAI Role, CAAI Group, Day Time, Night Time,
XC Flag, Is Simulator, Is Complex, Dual Instrument.

Logic matches caai_form_filler.py exactly for consistency.

Usage:
    python -m src.caai_columns --logbook logbook.xlsx
"""

import argparse
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter
from copy import copy

from .caai_rules import (
    is_simulator, get_caai_category, is_single_engine,
    normalize_type, is_complex_aircraft, CAAI_GROUP_MAP,
)
from .column_map import (
    COL_REGISTRATION, COL_AIRCRAFT_TYPE, COL_TOTAL_TIME, COL_PIC, COL_SIC,
    COL_NIGHT, COL_CROSS_COUNTRY, COL_ACTUAL_INSTRUMENT, COL_SIMULATED_INSTRUMENT,
    COL_DUAL_RECEIVED, COL_SOLO, COL_INSTRUCTOR, COL_DISTANCE, COL_REMARKS,
    COL_CAAI_ROLE, COL_CAAI_GROUP, COL_DAY_TIME, COL_NIGHT_TIME,
    COL_XC_FLAG, COL_IS_SIMULATOR, COL_IS_COMPLEX, COL_DUAL_INSTRUMENT,
    TOTAL_COLUMNS,
)

# ============ Styles ============
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
DATA_FONT = Font(name='Calibri', size=9)
THIN_BORDER = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
WRAP_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

# New column definitions: (header, width)
NEW_COLUMNS = [
    ("CAAI Role", 13),
    ("CAAI Group", 11),
    ("Day Time", 9),
    ("Night Time", 9),
    ("XC Flag", 8),
    ("Is Simulator", 11),
    ("Is Complex", 10),
    ("Dual Instrument", 13),
]


def add_caai_columns(logbook_file):
    """Add 8 CAAI classification columns (49-56) to the logbook.

    Args:
        logbook_file: Path to the flight logbook Excel file.

    Returns:
        Dict with categorization statistics.
    """
    wb = load_workbook(logbook_file)
    ws = wb["Flight Log"]

    # Write headers
    for i, (header_name, col_width) in enumerate(NEW_COLUMNS):
        col = COL_CAAI_ROLE + i
        cell = ws.cell(row=1, column=col, value=header_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = col_width

    # Tracking counters
    role_counts = Counter()
    role_hours = Counter()
    group_counts = Counter()
    group_hours = Counter()
    xc_count = 0
    sim_count = 0
    complex_count = 0
    dual_inst_count = 0

    # Process each flight row
    for row in range(2, ws.max_row + 1):
        aircraft_type = str(ws.cell(row=row, column=COL_AIRCRAFT_TYPE).value or '').strip()
        if not aircraft_type:
            continue

        registration = str(ws.cell(row=row, column=COL_REGISTRATION).value or '').strip()
        total = float(ws.cell(row=row, column=COL_TOTAL_TIME).value or 0)
        sic = float(ws.cell(row=row, column=COL_SIC).value or 0)
        night = float(ws.cell(row=row, column=COL_NIGHT).value or 0)
        xc = float(ws.cell(row=row, column=COL_CROSS_COUNTRY).value or 0)
        actual_inst = float(ws.cell(row=row, column=COL_ACTUAL_INSTRUMENT).value or 0)
        sim_inst = float(ws.cell(row=row, column=COL_SIMULATED_INSTRUMENT).value or 0)
        dual_recv = float(ws.cell(row=row, column=COL_DUAL_RECEIVED).value or 0)
        solo = float(ws.cell(row=row, column=COL_SOLO).value or 0)
        instructor = str(ws.cell(row=row, column=COL_INSTRUCTOR).value or '').strip()
        remarks = str(ws.cell(row=row, column=COL_REMARKS).value or '').strip()

        dist_val = ws.cell(row=row, column=COL_DISTANCE).value or 0
        if isinstance(dist_val, str):
            dist_val = dist_val.replace(',', '')
        distance = float(dist_val)

        # Get existing row fill
        existing_fill = ws.cell(row=row, column=2).fill
        row_fill = (existing_fill if existing_fill.start_color and
                    existing_fill.start_color.rgb and
                    existing_fill.start_color.rgb != '00000000' else None)

        sim = is_simulator(aircraft_type, registration)

        if sim:
            values = {
                COL_CAAI_ROLE: "N/A",
                COL_CAAI_GROUP: "SIM",
                COL_DAY_TIME: None,
                COL_NIGHT_TIME: None,
                COL_XC_FLAG: "No",
                COL_IS_SIMULATOR: "Yes",
                COL_IS_COMPLEX: "No",
                COL_DUAL_INSTRUMENT: "No",
            }
            sim_count += 1
            role_counts["N/A"] += 1
            group_counts["SIM"] += 1
            group_hours["SIM"] += total
        else:
            single_engine = is_single_engine(aircraft_type)
            has_instructor = bool(instructor) or dual_recv > 0
            is_safety = 'safety pilot' in remarks.lower()
            is_sic_field = sic > 0 and not has_instructor
            is_solo_flight = solo > 0
            is_xc = xc > 0 or distance > 27

            # CAAI Role (priority chain matching form filler exactly)
            if has_instructor:
                role = "Student"
            elif is_safety and single_engine:
                role = "Safety Pilot"
            elif is_sic_field and not single_engine:
                role = "SIC"
            elif is_sic_field and single_engine:
                role = "PIC"
            elif is_solo_flight:
                role = "PIC"
            else:
                role = "PIC"

            # CAAI Group
            caai_cat = get_caai_category(aircraft_type)
            group = CAAI_GROUP_MAP.get(caai_cat, 'א')

            # Day/Night
            day_time = round(total - night, 2)
            night_time = night

            # Complex
            is_complex = is_complex_aircraft(aircraft_type)

            # Dual Instrument
            has_inst_time = actual_inst > 0 or sim_inst > 0
            is_dual_inst = has_instructor and has_inst_time

            values = {
                COL_CAAI_ROLE: role,
                COL_CAAI_GROUP: group,
                COL_DAY_TIME: day_time if day_time > 0 else 0,
                COL_NIGHT_TIME: night_time if night_time > 0 else 0,
                COL_XC_FLAG: "Yes" if is_xc else "No",
                COL_IS_SIMULATOR: "No",
                COL_IS_COMPLEX: "Yes" if is_complex else "No",
                COL_DUAL_INSTRUMENT: "Yes" if is_dual_inst else "No",
            }

            role_counts[role] += 1
            role_hours[role] += total
            group_counts[group] += 1
            group_hours[group] += total
            if is_xc:
                xc_count += 1
            if is_complex:
                complex_count += 1
            if is_dual_inst:
                dual_inst_count += 1

        # Write values and apply formatting
        for col, value in values.items():
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            if col in (COL_DAY_TIME, COL_NIGHT_TIME) and isinstance(value, (int, float)):
                cell.number_format = '0.0'
            if row_fill:
                cell.fill = copy(row_fill)

    # Extend auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(TOTAL_COLUMNS)}1"

    wb.save(logbook_file)

    # Print verification
    print("=" * 60)
    print("CAAI COLUMNS ADDED SUCCESSFULLY")
    print("=" * 60)
    print(f"\nColumns {COL_CAAI_ROLE}-{COL_DUAL_INSTRUMENT} added to {ws.max_row - 1} flight rows")

    print(f"\n--- CAAI Role Counts ---")
    for role in ["PIC", "SIC", "Student", "Safety Pilot", "N/A"]:
        cnt = role_counts.get(role, 0)
        hrs = role_hours.get(role, 0)
        print(f"  {role:<15}: {cnt:>4} flights, {hrs:>8.1f} hrs")

    print(f"\n--- CAAI Group Counts ---")
    for group in ["א", "ב", "ג", "ד", "SIM"]:
        cnt = group_counts.get(group, 0)
        hrs = group_hours.get(group, 0)
        print(f"  {group:<4}: {cnt:>4} flights, {hrs:>8.1f} hrs")

    print(f"\n--- Flags ---")
    print(f"  XC flights: {xc_count}")
    print(f"  Simulators: {sim_count}")
    print(f"  Complex: {complex_count}")
    print(f"  Dual Instrument: {dual_inst_count}")

    return {
        'role_counts': dict(role_counts),
        'role_hours': dict(role_hours),
        'group_counts': dict(group_counts),
        'group_hours': dict(group_hours),
    }


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Add CAAI columns to flight logbook')
    parser.add_argument('--logbook', '-l', required=True, help='Flight logbook Excel file')
    args = parser.parse_args()
    add_caai_columns(args.logbook)
