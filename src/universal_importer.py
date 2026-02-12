"""
Universal flight logbook importer.

Reads flight data from any common format (Excel, CSV, TSV, PDF) and
produces a standardized 48-column Excel logbook compatible with the
CAAI processing pipeline.

Supports:
- Auto-format detection (by file extension / content sniffing)
- Auto-column detection (by header name matching)
- Explicit column mapping via config file
- Data normalization (dates, times, integers)

Usage:
    python -m src.universal_importer --input flights.xlsx --output logbook.xlsx
    python -m src.universal_importer --input flights.pdf --format pdf
    python -m src.universal_importer --input flights.csv --mapping my_mapping.ini
"""

import csv
import os
import re
import argparse
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font as OpenpyxlFont, PatternFill
from openpyxl.utils import get_column_letter

from .column_map import (
    COL_ROW_NUM, COL_DATE, COL_TOTAL_TIME, COL_PIC, COL_SIC, COL_NIGHT,
    COL_CROSS_COUNTRY, COL_ACTUAL_INSTRUMENT, COL_SIMULATED_INSTRUMENT,
    COL_DUAL_RECEIVED, COL_DUAL_GIVEN, COL_SOLO, COL_SIMULATOR,
    COL_MULTI_PILOT, COL_DAY_LANDINGS, COL_DAY_TAKEOFFS,
    COL_NIGHT_LANDINGS, COL_NIGHT_TAKEOFFS, COL_HOLDS, COL_GO_AROUNDS,
    COL_DISTANCE, COL_COMPLEX, COL_HIGH_PERF, COL_EFIS,
    COL_RETRACTABLE, COL_PRESSURIZED, COL_REMARKS,
    BASE_COLUMNS,
)
from .column_detector import (
    detect_columns, load_column_mapping, resolve_mapping_names,
    validate_mapping, print_mapping_report, COLUMN_NAMES,
)
from .logbook_creator import (
    HEADER_FONT, HEADER_FILL, DATA_FONT, THIN_BORDER, MAIN_COLUMNS,
)

# Try importing python-dateutil for robust date parsing
try:
    from dateutil import parser as dateutil_parser
    HAS_DATEUTIL = True
except ImportError:
    HAS_DATEUTIL = False

# Columns that contain time values (decimal hours)
TIME_COLUMNS = {
    COL_TOTAL_TIME, COL_PIC, COL_SIC, COL_NIGHT, COL_CROSS_COUNTRY,
    COL_ACTUAL_INSTRUMENT, COL_SIMULATED_INSTRUMENT,
    COL_DUAL_RECEIVED, COL_DUAL_GIVEN, COL_SOLO, COL_SIMULATOR,
    COL_MULTI_PILOT, COL_DISTANCE,
}

# Columns that contain integer values (counts)
INT_COLUMNS = {
    COL_DAY_LANDINGS, COL_DAY_TAKEOFFS, COL_NIGHT_LANDINGS,
    COL_NIGHT_TAKEOFFS, COL_HOLDS, COL_GO_AROUNDS,
}

# Columns that contain boolean-like values
BOOL_COLUMNS = {
    COL_COMPLEX, COL_HIGH_PERF, COL_EFIS, COL_RETRACTABLE, COL_PRESSURIZED,
}

# Columns that are pure text (no conversion)
TEXT_COLUMNS = set(range(1, BASE_COLUMNS + 1)) - TIME_COLUMNS - INT_COLUMNS - BOOL_COLUMNS - {COL_ROW_NUM, COL_DATE}


def detect_format(file_path):
    """Auto-detect file format from extension and content.

    Args:
        file_path: Path to the input file.

    Returns:
        Format string: 'excel', 'csv', 'tsv', 'pdf', or 'logten'.

    Raises:
        ValueError: If format cannot be determined.
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext in ('.xlsx', '.xls', '.xlsm'):
        return 'excel'
    elif ext == '.csv':
        return 'csv'
    elif ext == '.tsv':
        return 'tsv'
    elif ext == '.pdf':
        return 'pdf'
    elif ext == '.txt':
        # Check if it's a LogTen Pro export (tab-delimited with LogTen field names)
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                header = f.readline()
                if 'flight_flightDate' in header or 'flight_totalTime' in header:
                    return 'logten'
                elif '\t' in header:
                    return 'tsv'
                else:
                    return 'csv'
        except Exception:
            return 'csv'
    else:
        raise ValueError(
            f"Cannot determine format for '{file_path}' (extension: {ext}).\n"
            f"Supported formats: .xlsx, .csv, .tsv, .pdf, .txt"
        )


def _read_excel(file_path):
    """Read headers and data rows from an Excel file.

    Returns:
        Tuple of (headers: list[str], rows: list[list]).
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)

    # Use the first sheet (or 'Flight Log' if it exists)
    if 'Flight Log' in wb.sheetnames:
        ws = wb['Flight Log']
    else:
        ws = wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        raise ValueError(f"Excel file is empty: {file_path}")

    # First row = headers
    headers = [str(cell or '').strip() for cell in all_rows[0]]

    # Remaining rows = data
    data_rows = []
    for row in all_rows[1:]:
        # Skip empty rows
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        data_rows.append(list(row))

    print(f"  Read Excel: {len(data_rows)} rows, {len(headers)} columns")
    return headers, data_rows


def _read_csv(file_path, delimiter=','):
    """Read headers and data rows from a CSV/TSV file.

    Returns:
        Tuple of (headers: list[str], rows: list[list[str]]).
    """
    # Auto-detect delimiter if needed
    if delimiter is None:
        with open(file_path, 'r', encoding='utf-8') as f:
            sample = f.read(4096)
            sniffer = csv.Sniffer()
            try:
                dialect = sniffer.sniff(sample)
                delimiter = dialect.delimiter
            except csv.Error:
                delimiter = ','

    rows = []
    with open(file_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f, delimiter=delimiter)
        for row in reader:
            rows.append(row)

    if not rows:
        raise ValueError(f"File is empty: {file_path}")

    headers = [cell.strip() for cell in rows[0]]
    data_rows = []
    for row in rows[1:]:
        if all(not cell.strip() for cell in row):
            continue
        data_rows.append(row)

    fmt_name = 'TSV' if delimiter == '\t' else 'CSV'
    print(f"  Read {fmt_name}: {len(data_rows)} rows, {len(headers)} columns")
    return headers, data_rows


def _read_pdf(file_path):
    """Read headers and data rows from a PDF file.

    Returns:
        Tuple of (headers: list[str], rows: list[list[str]]).
    """
    from .pdf_reader import read_pdf_tables
    return read_pdf_tables(file_path)


def read_source(file_path, fmt='auto'):
    """Read raw data from any supported format.

    Args:
        file_path: Path to the source file.
        fmt: Format string ('auto', 'excel', 'csv', 'tsv', 'pdf').

    Returns:
        Tuple of (format_used: str, headers: list[str], rows: list[list]).
    """
    if fmt == 'auto':
        fmt = detect_format(file_path)

    print(f"  Format: {fmt}")

    if fmt == 'excel':
        headers, rows = _read_excel(file_path)
    elif fmt == 'csv':
        headers, rows = _read_csv(file_path, delimiter=',')
    elif fmt == 'tsv':
        headers, rows = _read_csv(file_path, delimiter='\t')
    elif fmt == 'pdf':
        headers, rows = _read_pdf(file_path)
    else:
        raise ValueError(f"Unsupported format: {fmt}")

    return fmt, headers, rows


def normalize_date(val):
    """Parse a date value in various formats.

    Handles: YYYY-MM-DD, DD/MM/YYYY, MM/DD/YYYY, DD-MM-YYYY,
    DD.MM.YYYY, datetime objects, and more via python-dateutil.

    Args:
        val: Date value (string, datetime, or None).

    Returns:
        datetime object, or None if parsing fails.
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return val

    s = str(val).strip()
    if not s:
        return None

    # Try common formats explicitly first
    for fmt in [
        '%Y-%m-%d',      # 2024-01-15
        '%d/%m/%Y',      # 15/01/2024 (Israeli/European)
        '%d-%m-%Y',      # 15-01-2024
        '%d.%m.%Y',      # 15.01.2024
        '%m/%d/%Y',      # 01/15/2024 (US)
        '%Y/%m/%d',      # 2024/01/15
        '%d %b %Y',      # 15 Jan 2024
        '%b %d, %Y',     # Jan 15, 2024
    ]:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue

    # Fall back to dateutil if available
    if HAS_DATEUTIL:
        try:
            # dayfirst=True for Israeli/European convention
            return dateutil_parser.parse(s, dayfirst=True)
        except (ValueError, TypeError):
            pass

    return None


def normalize_time(val):
    """Convert a time value to decimal hours.

    Handles: H:MM, HH:MM, decimal strings, floats, integers.

    Args:
        val: Time value (string, float, int, or None).

    Returns:
        Float decimal hours (rounded to 2 places), or None.
    """
    if val is None:
        return None

    if isinstance(val, (int, float)):
        return round(float(val), 2) if val != 0 else None

    s = str(val).strip()
    if not s or s == '0':
        return None

    # H:MM or HH:MM format
    match = re.match(r'^(\d+):(\d{1,2})$', s)
    if match:
        h, m = int(match.group(1)), int(match.group(2))
        return round(h + m / 60, 2)

    # Decimal with comma (European: "1,5" → 1.5)
    s = s.replace(',', '.')

    try:
        val = float(s)
        return round(val, 2) if val != 0 else None
    except ValueError:
        return None


def normalize_int(val):
    """Convert a value to integer (for landings, takeoffs, etc).

    Args:
        val: Value (string, float, int, or None).

    Returns:
        Integer, or 0 if conversion fails.
    """
    if val is None:
        return 0
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)

    s = str(val).strip()
    if not s:
        return 0

    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def create_standardized_logbook(input_file, output_file, pilot_name="",
                                 fmt="auto", column_mapping=None):
    """Universal import: any format -> standardized 48-column Excel logbook.

    This produces the same Excel format as logbook_creator.py, allowing
    the rest of the CAAI pipeline to work unchanged.

    Args:
        input_file: Path to the source file (Excel, CSV, TSV, or PDF).
        output_file: Path for the output Excel file.
        pilot_name: Pilot name for the summary sheet title.
        fmt: Format ('auto', 'excel', 'csv', 'tsv', 'pdf').
        column_mapping: Optional path to explicit column mapping file.

    Returns:
        Dict with summary statistics.
    """
    print(f"\nUniversal Import: {os.path.basename(input_file)}")
    print("=" * 70)

    # 1. Read source data
    fmt_used, headers, raw_rows = read_source(input_file, fmt)

    # 2. Detect or load column mapping
    if column_mapping and os.path.exists(column_mapping):
        print(f"  Loading explicit column mapping from: {column_mapping}")
        raw_mapping = load_column_mapping(column_mapping)
        mapping = resolve_mapping_names(raw_mapping, headers)
    else:
        print(f"  Auto-detecting column mapping...")
        mapping = detect_columns(headers)

    # 3. Print mapping report and validate
    print_mapping_report(mapping, headers)
    errors, warnings = validate_mapping(mapping, headers)

    if errors:
        print(f"\n  WARNING: {len(errors)} required columns not detected.")
        print(f"  The CAAI form may be incomplete. Consider providing a column mapping file.")
        print(f"  See docs/import-guide.md for instructions.\n")

    # 4. Create output workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Flight Log"

    # Write header row (same style as logbook_creator.py)
    for col_idx, (col_name, col_width) in enumerate(MAIN_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(MAIN_COLUMNS))}1"

    # 5. Write data rows with normalization
    flight_count = 0
    total_time_sum = 0.0
    skipped = 0

    for row_idx, raw_row in enumerate(raw_rows):
        # Extract values using mapping
        values = {}
        for our_col, src_idx in mapping.items():
            if src_idx < len(raw_row):
                values[our_col] = raw_row[src_idx]

        # Skip rows without a date or aircraft type (likely not flight records)
        date_val = normalize_date(values.get(COL_DATE))
        if date_val is None:
            skipped += 1
            continue

        flight_count += 1
        sheet_row = flight_count + 1  # +1 for header

        # Row number
        ws.cell(row=sheet_row, column=COL_ROW_NUM, value=flight_count).font = DATA_FONT
        ws.cell(row=sheet_row, column=COL_ROW_NUM).border = THIN_BORDER
        ws.cell(row=sheet_row, column=COL_ROW_NUM).alignment = Alignment(
            horizontal='center', vertical='center')

        # Date
        date_cell = ws.cell(row=sheet_row, column=COL_DATE, value=date_val)
        date_cell.font = DATA_FONT
        date_cell.border = THIN_BORDER
        date_cell.number_format = 'YYYY-MM-DD'
        date_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Process all other mapped columns
        for our_col, src_idx in mapping.items():
            if our_col in (COL_ROW_NUM, COL_DATE):
                continue

            raw_val = raw_row[src_idx] if src_idx < len(raw_row) else None

            # Normalize based on column type
            if our_col in TIME_COLUMNS:
                val = normalize_time(raw_val)
            elif our_col in INT_COLUMNS:
                val = normalize_int(raw_val)
                if val == 0:
                    val = None
            elif our_col == COL_DATE:
                continue  # Already handled above
            else:
                # Text column
                val = str(raw_val).strip() if raw_val is not None else None
                if val == '' or val == 'None':
                    val = None

            if val is not None:
                cell = ws.cell(row=sheet_row, column=our_col, value=val)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(
                    horizontal='center', vertical='center',
                    wrap_text=(our_col == COL_REMARKS))

                if our_col in TIME_COLUMNS and isinstance(val, (int, float)):
                    cell.number_format = '0.00'

        # Track totals
        total_val = normalize_time(values.get(COL_TOTAL_TIME))
        if total_val:
            total_time_sum += total_val

    # Add borders and formatting to empty cells for consistent grid
    for row in range(2, flight_count + 2):
        for col in range(1, len(MAIN_COLUMNS) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if cell.value is None:
                cell.font = DATA_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # ============ SHEET 2: SUMMARY ============
    ws2 = wb.create_sheet("Summary & Totals")

    title_font = OpenpyxlFont(name='Calibri', bold=True, size=14, color='1F4E79')
    section_font = OpenpyxlFont(name='Calibri', bold=True, size=12, color='FFFFFF')
    section_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    label_font = OpenpyxlFont(name='Calibri', bold=True, size=10)
    value_font = OpenpyxlFont(name='Calibri', size=10)

    for col_letter, width in [('A', 25), ('B', 15)]:
        ws2.column_dimensions[col_letter].width = width

    title = f"{pilot_name} - Flight Logbook Summary" if pilot_name else "Flight Logbook Summary"
    ws2.cell(row=1, column=1, value=title).font = title_font

    row = 3
    for col in range(1, 3):
        ws2.cell(row=row, column=col).fill = section_fill
    ws2.cell(row=row, column=1, value="IMPORT SUMMARY").font = section_font

    row = 4
    summary_items = [
        ("Source File", os.path.basename(input_file)),
        ("Format", fmt_used),
        ("Total Flights", flight_count),
        ("Total Time", f"{total_time_sum:.1f} hrs"),
        ("Columns Mapped", len(mapping)),
        ("Rows Skipped", skipped),
    ]

    for label, value in summary_items:
        ws2.cell(row=row, column=1, value=label).font = label_font
        ws2.cell(row=row, column=2, value=value).font = value_font
        ws2.cell(row=row, column=1).border = THIN_BORDER
        ws2.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    # Save
    wb.save(output_file)

    print(f"\n{'='*70}")
    print(f"Import complete!")
    print(f"  Source: {input_file} ({fmt_used})")
    print(f"  Flights: {flight_count}")
    print(f"  Total time: {total_time_sum:.1f} hrs")
    print(f"  Columns mapped: {len(mapping)}/{len(MAIN_COLUMNS)}")
    if skipped:
        print(f"  Rows skipped (no date): {skipped}")
    print(f"  Output: {output_file}")
    print(f"{'='*70}")

    return {
        'flights': flight_count,
        'total_time': total_time_sum,
        'columns_mapped': len(mapping),
        'format': fmt_used,
        'skipped': skipped,
    }


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Import any flight logbook format into standardized Excel')
    parser.add_argument('--input', '-i', required=True,
                        help='Source file (Excel, CSV, TSV, or PDF)')
    parser.add_argument('--output', '-o', required=True,
                        help='Output Excel file path')
    parser.add_argument('--format', '-f', default='auto',
                        choices=['auto', 'excel', 'csv', 'tsv', 'pdf'],
                        help='Source format (default: auto-detect)')
    parser.add_argument('--mapping', '-m', default=None,
                        help='Column mapping INI file (default: auto-detect)')
    parser.add_argument('--pilot', '-p', default='',
                        help='Pilot name for summary sheet')
    args = parser.parse_args()
    create_standardized_logbook(
        args.input, args.output, args.pilot, args.format, args.mapping)
