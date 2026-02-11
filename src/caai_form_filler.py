#!/usr/bin/env python3
"""
Fill the Israeli CAAI tofes-shaot (flight hours form) from the flight logbook.

Full CAAI compliance with all 8 rules:
1. Table 1 category totals = PIC + SIC + Student (no unaccounted hours)
   -> Safety pilot hours on SE aircraft EXCLUDED from category total
2. Student (מתלמד) = dual instruction only, cannot also be PIC
3. PIC = NOT safety pilot, NOT flights with instructor
4. PIC XC = NOT safety pilot, NOT flights with instructor
5. Actual instrument (not during instruction) on single-pilot aircraft = PIC
6. Simulator time NOT in Table 1 totals
7. Sim instrument in air during instruction = student time
8. SIC half-credit per תקנה 42(ב)

Usage:
    python -m src.caai_form_filler --logbook logbook.xlsx --template tofes-shaot.xlsx --output filled.xlsx
"""

import argparse
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.cell.cell import MergedCell
from collections import defaultdict

from .caai_rules import (
    is_simulator, get_caai_category, is_single_engine, normalize_type,
)
from .column_map import (
    COL_DATE, COL_FROM, COL_TO, COL_REGISTRATION, COL_AIRCRAFT_TYPE,
    COL_ENGINE_TYPE, COL_CLASS, COL_TOTAL_TIME, COL_PIC, COL_SIC,
    COL_NIGHT, COL_CROSS_COUNTRY, COL_ACTUAL_INSTRUMENT, COL_SIMULATED_INSTRUMENT,
    COL_DUAL_RECEIVED, COL_DUAL_GIVEN, COL_SOLO, COL_MULTI_PILOT,
    COL_DAY_LANDINGS, COL_NIGHT_LANDINGS, COL_INSTRUCTOR, COL_DISTANCE, COL_REMARKS,
)


DATA_FONT = Font(name='Calibri', size=9)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')


def set_cell(ws, row, col, value, fmt='0.0'):
    """Safely set a cell value, skipping merged cells."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value
    cell.font = DATA_FONT
    cell.alignment = CENTER_ALIGN
    if isinstance(value, (int, float)) and fmt:
        cell.number_format = fmt


def clear_cell(ws, row, col):
    """Clear a cell value."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = None


def read_logbook(logbook_file):
    """Read flight data from the logbook Excel file.

    Args:
        logbook_file: Path to the flight logbook Excel file.

    Returns:
        List of flight dicts.
    """
    wb = load_workbook(logbook_file)
    ws = wb["Flight Log"]

    flights = []
    for row in range(2, ws.max_row + 1):
        f = {}
        f['date'] = ws.cell(row=row, column=COL_DATE).value
        f['from'] = str(ws.cell(row=row, column=COL_FROM).value or '').strip()
        f['to'] = str(ws.cell(row=row, column=COL_TO).value or '').strip()
        f['reg'] = str(ws.cell(row=row, column=COL_REGISTRATION).value or '').strip()
        f['aircraft_type'] = str(ws.cell(row=row, column=COL_AIRCRAFT_TYPE).value or '').strip()
        f['engine'] = str(ws.cell(row=row, column=COL_ENGINE_TYPE).value or '').strip()
        f['class'] = str(ws.cell(row=row, column=COL_CLASS).value or '').strip()
        f['total'] = float(ws.cell(row=row, column=COL_TOTAL_TIME).value or 0)
        f['pic'] = float(ws.cell(row=row, column=COL_PIC).value or 0)
        f['sic'] = float(ws.cell(row=row, column=COL_SIC).value or 0)
        f['night'] = float(ws.cell(row=row, column=COL_NIGHT).value or 0)
        f['xc'] = float(ws.cell(row=row, column=COL_CROSS_COUNTRY).value or 0)
        f['actual_inst'] = float(ws.cell(row=row, column=COL_ACTUAL_INSTRUMENT).value or 0)
        f['sim_inst'] = float(ws.cell(row=row, column=COL_SIMULATED_INSTRUMENT).value or 0)
        f['dual_recv'] = float(ws.cell(row=row, column=COL_DUAL_RECEIVED).value or 0)
        f['dual_given'] = float(ws.cell(row=row, column=COL_DUAL_GIVEN).value or 0)
        f['solo'] = float(ws.cell(row=row, column=COL_SOLO).value or 0)
        f['multi_pilot'] = float(ws.cell(row=row, column=COL_MULTI_PILOT).value or 0)
        f['day_ldg'] = int(ws.cell(row=row, column=COL_DAY_LANDINGS).value or 0)
        f['night_ldg'] = int(ws.cell(row=row, column=COL_NIGHT_LANDINGS).value or 0)
        f['instructor'] = str(ws.cell(row=row, column=COL_INSTRUCTOR).value or '').strip()
        f['remarks'] = str(ws.cell(row=row, column=COL_REMARKS).value or '').strip()
        dist_val = ws.cell(row=row, column=COL_DISTANCE).value or 0
        if isinstance(dist_val, str):
            dist_val = dist_val.replace(',', '')
        f['distance'] = float(dist_val)
        flights.append(f)

    wb.close()
    return flights


def categorize_flights(flights):
    """Categorize all flights per CAAI rules.

    Args:
        flights: List of flight dicts from read_logbook().

    Returns:
        Tuple of (type_stats dict, grand totals dict, special flight lists).
    """
    type_stats = defaultdict(lambda: {
        'caai_col': 'C', 'is_sim': False,
        'total': 0, 'form_total': 0,
        'day_pic': 0, 'day_pic_xc': 0, 'day_sic': 0, 'day_student': 0,
        'night_pic': 0, 'night_pic_xc': 0, 'night_sic': 0, 'night_student': 0,
        'inst_actual': 0, 'inst_sim_air': 0, 'inst_sim_device': 0,
        'dual_recv': 0, 'dual_inst': 0, 'night_time': 0, 'solo': 0,
        'night_ldg': 0, 'complex_time': 0, 'safety_pilot': 0,
    })

    grand = {
        'pic': 0, 'pic_xc': 0, 'sic': 0, 'student': 0,
        'night_pic': 0, 'night_pic_xc': 0, 'night_sic': 0, 'night_student': 0,
        'actual_inst': 0, 'sim_inst_air': 0, 'sim_device': 0,
        'total': 0, 'form_total': 0,
        'night': 0, 'dual': 0, 'dual_inst': 0, 'solo': 0, 'solo_xc': 0,
        'night_ldg': 0, 'xc': 0, 'xc_all_roles': 0,
        'complex': 0, 'safety_pilot_se': 0, 'me_time': 0,
    }

    night_pic_flights = []
    inst_instruction_flights = []
    complex_flights = []

    for f in flights:
        if not f['aircraft_type']:
            continue

        if is_simulator(f['aircraft_type'], f['reg']):
            ntype = normalize_type(f['aircraft_type'])
            type_stats[ntype]['is_sim'] = True
            type_stats[ntype]['inst_sim_device'] += f['total']
            grand['sim_device'] += f['total']
            continue

        ntype = normalize_type(f['aircraft_type'])
        ts = type_stats[ntype]
        ts['caai_col'] = get_caai_category(f['aircraft_type'])
        ts['total'] += f['total']

        day_time = f['total'] - f['night']
        night_time = f['night']
        single_engine = is_single_engine(f['aircraft_type'])

        has_instructor = bool(f['instructor']) or f['dual_recv'] > 0
        is_safety = 'safety pilot' in f['remarks'].lower()
        is_sic_field = f['sic'] > 0 and not has_instructor
        is_solo_flight = f['solo'] > 0
        is_xc = f['xc'] > 0 or f['distance'] > 27

        # Instrument time (always counted)
        ts['inst_actual'] += f['actual_inst']
        ts['inst_sim_air'] += f['sim_inst']
        grand['actual_inst'] += f['actual_inst']
        grand['sim_inst_air'] += f['sim_inst']

        ts['night_time'] += night_time
        ts['night_ldg'] += f['night_ldg']
        grand['total'] += f['total']
        grand['night'] += night_time
        grand['night_ldg'] += f['night_ldg']

        if is_xc:
            grand['xc_all_roles'] += f['total']

        if not single_engine:
            grand['me_time'] += f['total']

        # Complex: PA44, BE76
        if ntype in ['PA44', 'BE76']:
            ts['complex_time'] += f['total']
            grand['complex'] += f['total']
            if f['total'] > 0:
                complex_flights.append(f)

        # === ROLE CLASSIFICATION ===
        if has_instructor:
            ts['day_student'] += day_time
            ts['night_student'] += night_time
            ts['form_total'] += f['total']
            ts['dual_recv'] += f['total']
            grand['student'] += f['total']
            grand['dual'] += f['total']
            grand['form_total'] += f['total']
            if night_time > 0:
                grand['night_student'] += night_time
            if f['actual_inst'] > 0 or f['sim_inst'] > 0:
                ts['dual_inst'] += f['actual_inst'] + f['sim_inst']
                grand['dual_inst'] += f['actual_inst'] + f['sim_inst']
                inst_instruction_flights.append(f)
        elif is_safety and single_engine:
            ts['safety_pilot'] += f['total']
            grand['safety_pilot_se'] += f['total']
        elif is_sic_field and not single_engine:
            ts['day_sic'] += day_time
            ts['night_sic'] += night_time
            ts['form_total'] += f['total']
            grand['sic'] += f['total']
            grand['form_total'] += f['total']
            grand['night_sic'] += night_time
        elif is_sic_field and single_engine:
            ts['day_pic'] += day_time
            ts['night_pic'] += night_time
            ts['form_total'] += f['total']
            grand['pic'] += f['total']
            grand['form_total'] += f['total']
            if is_xc:
                ts['day_pic_xc'] += day_time
                ts['night_pic_xc'] += night_time
                grand['pic_xc'] += f['total']
                grand['xc'] += f['total']
            if night_time > 0:
                grand['night_pic'] += night_time
                if is_xc:
                    grand['night_pic_xc'] += night_time
                night_pic_flights.append(f)
        elif is_solo_flight:
            ts['day_pic'] += day_time
            ts['night_pic'] += night_time
            ts['form_total'] += f['total']
            ts['solo'] += f['total']
            grand['pic'] += f['total']
            grand['form_total'] += f['total']
            grand['solo'] += f['total']
            if is_xc:
                ts['day_pic_xc'] += day_time
                ts['night_pic_xc'] += night_time
                grand['pic_xc'] += f['total']
                grand['solo_xc'] += f['total']
                grand['xc'] += f['total']
            if night_time > 0:
                grand['night_pic'] += night_time
                if is_xc:
                    grand['night_pic_xc'] += night_time
                night_pic_flights.append(f)
        else:
            ts['day_pic'] += day_time
            ts['night_pic'] += night_time
            ts['form_total'] += f['total']
            grand['pic'] += f['total']
            grand['form_total'] += f['total']
            if is_xc:
                ts['day_pic_xc'] += day_time
                ts['night_pic_xc'] += night_time
                grand['pic_xc'] += f['total']
                grand['xc'] += f['total']
            if night_time > 0:
                grand['night_pic'] += night_time
                if is_xc:
                    grand['night_pic_xc'] += night_time
                night_pic_flights.append(f)

    return type_stats, grand, {
        'night_pic_flights': night_pic_flights,
        'inst_instruction_flights': inst_instruction_flights,
        'complex_flights': complex_flights,
    }


def fill_caai_form(logbook_file, template_file, output_file):
    """Fill the CAAI tofes-shaot form from logbook data.

    Copies the template, then fills it with categorized flight data.

    Args:
        logbook_file: Path to the flight logbook Excel file.
        template_file: Path to the blank CAAI form template.
        output_file: Path for the filled output form.

    Returns:
        Dict with verification statistics.
    """
    # Copy template to output (don't modify template)
    shutil.copy2(template_file, output_file)

    # Read and categorize flights
    flights = read_logbook(logbook_file)
    type_stats, grand, special = categorize_flights(flights)

    # Print categorization
    print("=" * 90)
    print("CAAI CATEGORIZATION (FULL COMPLIANCE)")
    print("=" * 90)
    for atype in sorted(type_stats.keys(), key=lambda x: type_stats[x]['total'], reverse=True):
        t = type_stats[atype]
        if t['is_sim']:
            print(f"  {atype:<12} SIM  device={t['inst_sim_device']:.1f}")
        elif t['total'] > 0:
            safety = f" safety={t['safety_pilot']:.1f}" if t['safety_pilot'] > 0 else ""
            print(f"  {atype:<12} {t['caai_col']}  total={t['total']:.1f}  form={t['form_total']:.1f}  dPIC={t['day_pic']:.1f}  dSIC={t['day_sic']:.1f}  dSTD={t['day_student']:.1f}  nPIC={t['night_pic']:.1f}  nSIC={t['night_sic']:.1f}  nSTD={t['night_student']:.1f}{safety}")

    print(f"\nGrand totals:")
    print(f"  Total aircraft hours: {grand['total']:.1f}")
    print(f"  Form total (excl safety): {grand['form_total']:.1f}")
    print(f"  PIC: {grand['pic']:.1f}")
    print(f"  SIC: {grand['sic']:.1f}")
    print(f"  Student: {grand['student']:.1f}")
    print(f"  Safety pilot SE: {grand['safety_pilot_se']:.1f}")
    print(f"  Form check: PIC+SIC+Student = {grand['pic']+grand['sic']+grand['student']:.1f} vs form_total = {grand['form_total']:.1f}")

    # Fill the form
    wb = load_workbook(output_file)

    # === Summary Sheet ===
    ws = wb['סיכום ניסיון תעופתי']

    # Clear old data
    for row in range(13, 23):
        for col in range(2, 21):
            clear_cell(ws, row, col)
    for row in range(31, 41):
        for col in range(3, 7):
            clear_cell(ws, row, col)

    col_map = {'C': 3, 'D': 4, 'E': 5, 'F': 6}

    real_types = [(k, v) for k, v in type_stats.items() if not v['is_sim'] and v['total'] > 0]
    real_types.sort(key=lambda x: x[1]['form_total'], reverse=True)
    sim_types = [(k, v) for k, v in type_stats.items() if v['is_sim']]

    print(f"\n{'='*60}")
    print(f"FILLING SUMMARY SHEET - {len(real_types)} aircraft types")
    print(f"{'='*60}")

    for i, (atype, ts) in enumerate(real_types):
        if i >= 10:
            print(f"WARNING: More than 10 types, truncating!")
            break
        form_row = 13 + i
        caai_col = ts['caai_col']

        set_cell(ws, form_row, 2, atype, None)
        set_cell(ws, form_row, col_map[caai_col], round(ts['form_total'], 1))

        if ts['day_pic'] > 0: set_cell(ws, form_row, 13, round(ts['day_pic'], 1))
        if ts['day_pic_xc'] > 0: set_cell(ws, form_row, 14, round(ts['day_pic_xc'], 1))
        if ts['day_sic'] > 0: set_cell(ws, form_row, 15, round(ts['day_sic'], 1))
        if ts['day_student'] > 0: set_cell(ws, form_row, 16, round(ts['day_student'], 1))
        if ts['night_pic'] > 0: set_cell(ws, form_row, 17, round(ts['night_pic'], 1))
        if ts['night_pic_xc'] > 0: set_cell(ws, form_row, 18, round(ts['night_pic_xc'], 1))
        if ts['night_sic'] > 0: set_cell(ws, form_row, 19, round(ts['night_sic'], 1))
        if ts['night_student'] > 0: set_cell(ws, form_row, 20, round(ts['night_student'], 1))

        role_sum = (ts['day_pic'] + ts['day_sic'] + ts['day_student'] +
                    ts['night_pic'] + ts['night_sic'] + ts['night_student'])
        diff = abs(ts['form_total'] - role_sum)
        status = "OK" if diff < 0.2 else f"MISMATCH={diff:.1f}"
        print(f"  Row {form_row}: {atype:<8} ({caai_col}) form_total={ts['form_total']:.1f}  roles={role_sum:.1f}  [{status}]")

    # Table 2 - Instrument time
    print(f"\nTable 2 - Instrument:")
    sim_row_map = {}
    for i, (atype, ts) in enumerate(real_types[:10]):
        form_row = 31 + i
        sim_row_map[atype] = form_row
        if ts['inst_actual'] > 0: set_cell(ws, form_row, 3, round(ts['inst_actual'], 1))
        if ts['inst_sim_air'] > 0: set_cell(ws, form_row, 4, round(ts['inst_sim_air'], 1))

    for stype, sts in sim_types:
        base_type = stype.replace(' SIM', '').replace(' FTD', '').replace(' FFS', '')
        if base_type == 'FRASCA': base_type = 'C172'
        elif base_type == 'A320': base_type = 'A319'
        elif base_type == 'FLIGHT SAFETY': base_type = 'H25B'
        elif base_type == 'ATP - CTP TRAINING': base_type = 'A319'

        if base_type in sim_row_map:
            form_row = sim_row_map[base_type]
            existing = ws.cell(row=form_row, column=5).value or 0
            new_val = round(existing + sts['inst_sim_device'], 1)
            set_cell(ws, form_row, 5, new_val)
            print(f"  {stype} -> {base_type} device={sts['inst_sim_device']:.1f}")

    # === CPL Sheet ===
    ws_cpl = wb['רישיון טיס מסחרי']

    for row in range(12, 19):
        for col in range(3, 7):
            clear_cell(ws_cpl, row, col)
    for row in range(26, 50):
        for col in range(2, 7):
            clear_cell(ws_cpl, row, col)
    for col in [8, 11, 14]:
        clear_cell(ws_cpl, 17, col)

    print(f"\n{'='*60}")
    print(f"CPL SHEET")
    print(f"{'='*60}")

    set_cell(ws_cpl, 12, 3, round(grand['pic_xc'], 1))
    print(f"  C12 PIC XC: {grand['pic_xc']:.1f}")

    set_cell(ws_cpl, 13, 3, round(grand['dual'], 1))
    print(f"  C13 Dual received: {grand['dual']:.1f}")

    set_cell(ws_cpl, 14, 3, round(grand['dual_inst'], 1))
    print(f"  C14 Dual instrument: {grand['dual_inst']:.1f}")

    set_cell(ws_cpl, 15, 3, grand['night_ldg'], '0')
    print(f"  C15 Night landings: {grand['night_ldg']}")

    set_cell(ws_cpl, 16, 3, round(grand['night'], 1))
    print(f"  C16 Night hours: {grand['night']:.1f}")

    # C17 = Solo XC long flight
    solo_xc_flights = [f for f in flights if f['solo'] > 0 and (f['xc'] > 0 or f['distance'] > 27)
                       and not is_simulator(f['aircraft_type'], f['reg'])]
    if solo_xc_flights:
        longest_solo = max(solo_xc_flights, key=lambda x: x['distance'])
        set_cell(ws_cpl, 17, 3, round(longest_solo['total'], 1))
        if longest_solo['date']:
            date_str = (longest_solo['date'].strftime('%d/%m/%Y')
                        if hasattr(longest_solo['date'], 'strftime')
                        else str(longest_solo['date']))
            set_cell(ws_cpl, 17, 8, date_str, None)
        dist_km = round(longest_solo['distance'] * 1.852, 0)
        set_cell(ws_cpl, 17, 11, int(dist_km), '0')
        route = f"{longest_solo['from']}-{longest_solo['to']}"
        set_cell(ws_cpl, 17, 14, route, None)
        print(f"  C17 Solo XC long: {longest_solo['total']:.1f} hrs, {route}, {dist_km:.0f}km")

    # C18 = Complex/Group B+C
    complex_or_group_bc = grand['complex']
    for atype, ts in type_stats.items():
        if ts['caai_col'] == 'F' and not ts['is_sim']:
            complex_or_group_bc += ts['form_total']
    set_cell(ws_cpl, 18, 3, round(complex_or_group_bc, 1))
    print(f"  C18 Complex/Group B+C: {complex_or_group_bc:.1f}")

    # CPL Table 2
    print(f"\n  CPL Table 2:")
    sorted_inst = sorted(special['inst_instruction_flights'], key=lambda x: str(x['date']))
    print(f"  Instrument instruction: {len(sorted_inst)} flights")
    for i, f in enumerate(sorted_inst[:20]):
        r = 27 + i
        date_str = f['date'].strftime('%d/%m/%Y') if hasattr(f['date'], 'strftime') else str(f['date'])
        inst_time = round(f['actual_inst'] + f['sim_inst'], 1)
        set_cell(ws_cpl, r, 2, f"{date_str}  {inst_time:.1f}", None)

    sorted_night = sorted(special['night_pic_flights'], key=lambda x: str(x['date']))
    print(f"  Night PIC: {len(sorted_night)} flights")
    for i, f in enumerate(sorted_night[:20]):
        r = 27 + i
        date_str = f['date'].strftime('%d/%m/%Y') if hasattr(f['date'], 'strftime') else str(f['date'])
        set_cell(ws_cpl, r, 3, date_str, None)
        set_cell(ws_cpl, r, 4, round(f['night'], 1))

    sorted_complex = sorted(special['complex_flights'], key=lambda x: str(x['date']))
    print(f"  Complex: {len(sorted_complex)} flights")
    for i, f in enumerate(sorted_complex[:20]):
        r = 27 + i
        date_str = f['date'].strftime('%d/%m/%Y') if hasattr(f['date'], 'strftime') else str(f['date'])
        set_cell(ws_cpl, r, 5, date_str, None)
        set_cell(ws_cpl, r, 6, round(f['total'], 1))

    # === ATPL Sheet ===
    ws_atpl = wb['רישיון טיס תובלה בנתיבי אוויר']

    print(f"\n{'='*60}")
    print(f"ATPL SHEET")
    print(f"{'='*60}")

    set_cell(ws_atpl, 13, 3, round(grand['xc_all_roles'], 1))
    print(f"  C13 XC total (all roles): {grand['xc_all_roles']:.1f}")

    set_cell(ws_atpl, 14, 3, round(grand['night_pic_xc'], 1))
    print(f"  C14 Night PIC XC: {grand['night_pic_xc']:.1f}")

    total_inst_aircraft = grand['actual_inst'] + grand['sim_inst_air']
    set_cell(ws_atpl, 15, 3, round(total_inst_aircraft, 1))
    print(f"  C15 Instrument: {total_inst_aircraft:.1f}")

    # Save
    wb.save(output_file)
    print(f"\nForm saved: {output_file}")

    # Verification
    print(f"\n{'='*60}")
    print("CAAI COMPLIANCE VERIFICATION")
    print(f"{'='*60}")

    print(f"\n--- Table 1 ---")
    print(f"  Safety pilot excluded: {grand['safety_pilot_se']:.1f} hrs")
    print(f"  Form total (PIC+SIC+Student): {grand['form_total']:.1f}")
    sum_check = grand['pic'] + grand['sic'] + grand['student']
    diff = abs(grand['form_total'] - sum_check)
    print(f"  Sum check: {sum_check:.1f} (diff={diff:.1f}) {'OK' if diff < 0.5 else 'INVESTIGATE'}")

    print(f"\n--- Grand Total (with SIC half-credit) ---")
    caai_grand = grand['pic'] + grand['sic'] / 2 + grand['student']
    print(f"  PIC={grand['pic']:.1f} + SIC/2={grand['sic']/2:.1f} + Student={grand['student']:.1f} = {caai_grand:.1f}")

    return {
        'grand': grand,
        'type_stats': dict(type_stats),
        'caai_grand_total': caai_grand,
    }


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Fill CAAI tofes-shaot form')
    parser.add_argument('--logbook', '-l', required=True, help='Flight logbook Excel file')
    parser.add_argument('--template', '-t', required=True, help='Blank CAAI form template')
    parser.add_argument('--output', '-o', required=True, help='Output filled form file')
    args = parser.parse_args()
    fill_caai_form(args.logbook, args.template, args.output)
